# -*- coding: utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import time
import traceback
import sys
import os
from pathlib import Path
import shutil
from datetime import datetime, timedelta
import logging # <-- MODIFICA: Aggiunto modulo logging

# --- CONFIGURAZIONE LOGGING (AGGIUNTO) ---
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)-8s - %(message)s", handlers=[logging.StreamHandler()])
logger = logging.getLogger(__name__)

# --- FUNZIONE DI ATTESA ROBUSTA (AGGIUNTA) ---
def attendi_scomparsa_overlay(driver, timeout_secondi=45):
    """
    Attende in modo robusto che gli overlay di caricamento tipici dei siti Ext JS scompaiano.
    Rileva maschere di caricamento o messaggi specifici come 'Caricamento...'.
    """
    try:
        overlay_wait = WebDriverWait(driver, timeout_secondi)
        # XPath combinato per maschere di caricamento comuni in Ext JS o testo di caricamento
        xpath_overlay = "//div[contains(@class, 'x-mask-msg') or contains(@class, 'x-mask')][not(contains(@style,'display: none'))] | //div[text()='Caricamento...']"
        
        # Attende che gli elementi di cui sopra diventino invisibili
        overlay_wait.until(EC.invisibility_of_element_located((By.XPATH, xpath_overlay)))
        logger.info(" -> Overlay di caricamento scomparso.")
        time.sleep(0.3) # Piccola pausa per stabilizzare l'interfaccia
        return True
    except TimeoutException:
        logger.warning(f"Timeout ({timeout_secondi}s) durante l'attesa della scomparsa dell'overlay. Proseguo con cautela.")
        return False

# --- Variabili Globali Configurabili ---
LOGIN_URL = "https://portalefornitori.isab.com/Ui/"
FORNITORE_DA_SELEZIONARE = "KK10608 - COEMI S.R.L."

# --- Dettagli File di Configurazione e Database ---
SCRIPT_DIRECTORY = Path(__file__).resolve().parent
CONFIG_EXCEL_FILENAME = "parametriScaricoTS.xlsm"
CONFIG_EXCEL_PATH = SCRIPT_DIRECTORY / CONFIG_EXCEL_FILENAME

DATABASE_FILENAME = "database_timbrature_isab.xlsm"
DATABASE_FILE_PATH = SCRIPT_DIRECTORY / DATABASE_FILENAME
DATABASE_SHEET_NAME = "Dati"

# --- Celle per la Configurazione ---
SHEET_NAME_CONFIG = "parametri"
USERNAME_CELL = "A3"
PASSWORD_CELL = "B3"
DOWNLOAD_DIR_CELL = "E2"

# --- Inizializzazione Variabili di Configurazione ---
USERNAME = None
PASSWORD = None
DOWNLOAD_DIR = None

# --- Sezione 1: Lettura della Configurazione ---
logger.info(f"Tentativo di leggere i dati di configurazione dal file: {CONFIG_EXCEL_PATH}")
try:
    if not CONFIG_EXCEL_PATH.exists():
        logger.critical(f"Errore FATALE: File Excel di configurazione non trovato: {CONFIG_EXCEL_PATH}")
        sys.exit("File Excel di configurazione mancante.")

    workbook_config = openpyxl.load_workbook(CONFIG_EXCEL_PATH, data_only=True)
    if SHEET_NAME_CONFIG in workbook_config.sheetnames:
        sheet_config = workbook_config[SHEET_NAME_CONFIG]
        USERNAME = sheet_config[USERNAME_CELL].value
        PASSWORD = sheet_config[PASSWORD_CELL].value
        download_dir_excel = sheet_config[DOWNLOAD_DIR_CELL].value
        if download_dir_excel and isinstance(download_dir_excel, str) and download_dir_excel.strip():
            DOWNLOAD_DIR = download_dir_excel.strip()
        else:
            logger.critical(f"Errore FATALE: Percorso DOWNLOAD_DIR non valido o vuoto nella cella {DOWNLOAD_DIR_CELL}.")
            sys.exit("Percorso download mancante.")
        if not (USERNAME and PASSWORD):
            logger.critical("Errore FATALE: Username o Password non trovati nel file Excel.")
            sys.exit("Credenziali mancanti.")
        logger.info("Configurazione letta con successo.")
        logger.info(f"  DOWNLOAD_DIR: '{DOWNLOAD_DIR}'")
    else:
        logger.critical(f"Errore FATALE: Foglio '{SHEET_NAME_CONFIG}' non trovato nel file Excel.")
        sys.exit(f"Foglio '{SHEET_NAME_CONFIG}' non trovato.")
    workbook_config.close()

except Exception as e_excel_config:
    logger.critical(f"Errore FATALE durante la lettura della configurazione da Excel: {e_excel_config}\n{traceback.format_exc()}")
    sys.exit("Errore configurazione Excel.")


# --- Sezione 2: Automazione Web con Selenium ---
logger.info("\nAvvio script automatico per operazioni web...")
driver = None
final_downloaded_path = None

try:
    yesterday = datetime.now() - timedelta(days=1)
    data_da_usare = yesterday.strftime('%d.%m.%Y')
    logger.info(f"Data calcolata per la ricerca (ieri): {data_da_usare}")

    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": str(Path(DOWNLOAD_DIR).absolute()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True, 
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_settings.popups": 0,
        "profile.content_settings.exceptions.automatic_downloads.*.setting": 1
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Argomenti per disabilitare le nuove feature di sicurezza di Chrome (Bubble, Warnings)
    chrome_options.add_argument("--disable-features=InsecureDownloadWarnings")
    chrome_options.add_argument("--disable-features=DownloadBubble,DownloadBubbleV2")

    # Altri argomenti permissivi
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument(f"--unsafely-treat-insecure-origin-as-secure={LOGIN_URL}")

    logger.info("Inizializzazione WebDriver Chrome...")
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 20)
    long_wait = WebDriverWait(driver, 30)
    popup_wait = WebDriverWait(driver, 7)

    logger.info(f"Navigazione a: {LOGIN_URL}")
    driver.get(LOGIN_URL)

    logger.info("Tentativo di login...")
    wait.until(EC.presence_of_element_located((By.NAME, "Username"))).send_keys(USERNAME)
    wait.until(EC.presence_of_element_located((By.NAME, "Password"))).send_keys(PASSWORD)
    
    logger.info("Click sul pulsante 'Accedi'...")
    accedi_button_xpath = "//span[text()='Accedi' and contains(@class, 'x-btn-inner')]"
    wait.until(EC.element_to_be_clickable((By.XPATH, accedi_button_xpath))).click()
    
    logger.info("Verifica del successo del login in corso...")
    attendi_scomparsa_overlay(driver, 60) # <-- MODIFICA: attesa robusta post-login

    # La verifica post-login rimane un'ottima pratica
    try:
        report_button_xpath_verification = "//*[normalize-space(text())='Report']"
        wait.until(EC.visibility_of_element_located((By.XPATH, report_button_xpath_verification)))
        logger.info("Login effettuato con successo. Elemento 'Report' visibile.")
    except TimeoutException:
        logger.critical("\n" + "="*60)
        logger.critical("ERRORE FATALE: LOGIN FALLITO.")
        logger.critical("L'elemento 'Report' non è stato trovato dopo il login.")
        logger.critical("Possibili cause: credenziali errate, CAPTCHA, lentezza del sito.")
        logger.critical("="*60)
        # Salva screenshot in caso di fallimento
        screenshot_path = SCRIPT_DIRECTORY / f"login_failure_{datetime.now():%Y%m%d_%H%M%S}.png"
        try:
            driver.save_screenshot(str(screenshot_path))
            logger.info(f"Screenshot del fallimento salvato in: {screenshot_path}")
        except Exception as e_shot:
            logger.error(f"Impossibile salvare lo screenshot: {e_shot}")
        sys.exit("Login non riuscito. Script interrotto.")

    try:
        ok_button_popup = popup_wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='OK' and contains(@class, 'x-btn-inner')]")))
        ok_button_popup.click()
        logger.info("Pop-up 'OK' trovato e cliccato.")
    except TimeoutException:
        logger.info("Nessun pop-up 'OK' rilevato (normale).")

    logger.info("Navigazione menu: Report -> Timbrature")
    wait.until(EC.element_to_be_clickable((By.XPATH, "//*[normalize-space(text())='Report']"))).click()
    attendi_scomparsa_overlay(driver) # <-- MODIFICA
    
    wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Timbrature']"))).click()
    logger.info("'Timbrature' cliccato.")
    attendi_scomparsa_overlay(driver) # <-- MODIFICA

    logger.info("Impostazione filtri per il report Timbrature...")
    fornitore_arrow_xpath = "//div[starts-with(@id, 'generic_refresh_combo_box-') and contains(@id, '-trigger-picker')]"
    wait.until(EC.visibility_of_element_located((By.XPATH, fornitore_arrow_xpath)))

    logger.info(f"  Seleziono il fornitore: '{FORNITORE_DA_SELEZIONARE}'...")
    ActionChains(driver).move_to_element(wait.until(EC.element_to_be_clickable((By.XPATH, fornitore_arrow_xpath)))).click().perform()
    attendi_scomparsa_overlay(driver) # <-- MODIFICA: attende il caricamento del dropdown
    
    fornitore_option_xpath = f"//li[normalize-space(text())='{FORNITORE_DA_SELEZIONARE}']"
    fornitore_option = long_wait.until(EC.presence_of_element_located((By.XPATH, fornitore_option_xpath)))
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", fornitore_option)
    logger.info(f"  Fornitore '{FORNITORE_DA_SELEZIONARE}' selezionato.")
    attendi_scomparsa_overlay(driver) # <-- MODIFICA: attende l'aggiornamento post-selezione

    # Verifica se il campo è stato effettivamente popolato
    try:
        valore_input = driver.find_element(By.NAME, "CodiceFornitore").get_attribute("value")
        if not valore_input:
            logger.warning("  ATTENZIONE: Il campo Fornitore risulta vuoto! Provo inserimento manuale...")
            campo_f = driver.find_element(By.NAME, "CodiceFornitore")
            campo_f.send_keys(FORNITORE_DA_SELEZIONARE)
            time.sleep(1)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo_f)
    except Exception as e_verify:
        logger.debug(f"  Errore durante la verifica del campo fornitore: {e_verify}")

    logger.info(f"  Inserimento data Da/A: '{data_da_usare}'...")
    wait.until(EC.visibility_of_element_located((By.NAME, "DataTsDa"))).clear()
    driver.find_element(By.NAME, "DataTsDa").send_keys(data_da_usare)
    wait.until(EC.visibility_of_element_located((By.NAME, "DataTsA"))).clear()
    driver.find_element(By.NAME, "DataTsA").send_keys(data_da_usare)
    logger.info("  Date inserite.")
    
    logger.info("  Click sul pulsante 'Cerca'...")
    cerca_button_xpath = "//a[contains(@class, 'x-btn') and .//span[normalize-space(text())='Cerca']]"
    wait.until(EC.element_to_be_clickable((By.XPATH, cerca_button_xpath))).click()
    logger.info("  Pulsante 'Cerca' cliccato. Attesa risultati...")
    attendi_scomparsa_overlay(driver, 90) # <-- MODIFICA: attesa lunga per i risultati della ricerca

    logger.info("  Tentativo di download del file Excel...")
    path_to_downloads_obj = Path(DOWNLOAD_DIR)
    files_before_download = set(path_to_downloads_obj.iterdir())
    
    excel_button_xpath = "//div[contains(@class, 'x-tool') and @role='button' and .//div[contains(@style, 'FontAwesome')]]"
    excel_button = wait.until(EC.element_to_be_clickable((By.XPATH, excel_button_xpath)))
    
    logger.info("  Icona Excel per download trovata. Clicco via JS...")
    driver.execute_script("arguments[0].click();", excel_button)
    logger.info("  Click eseguito. Attendo il completamento (max 45s)...")
    
    download_start_time = time.time()
    while time.time() - download_start_time < 45:
        current_files = set(path_to_downloads_obj.iterdir())
        new_files = current_files - files_before_download
        completed_files = [f for f in new_files if f.suffix.lower() == '.xlsx' and not f.name.endswith(('.crdownload', '.tmp'))]
        if completed_files:
            final_downloaded_path = max(completed_files, key=lambda f: f.stat().st_mtime)
            time.sleep(1) # Pausa per il flush del file system
            if final_downloaded_path.exists() and final_downloaded_path.stat().st_size > 0:
                logger.info(f"  Download COMPLETATO. File rilevato: {final_downloaded_path.name}")
                break
            else:
                final_downloaded_path = None # Resetta se il file non è valido
        time.sleep(1) 
    
    if not final_downloaded_path:
        logger.critical("ERRORE CRITICO: Download del report timbrature fallito o file non valido.")
    
    logger.info("-" * 40)
    logger.info("Tentativo di Logout...")
    try:
        settings_button_xpath = "//span[@id='user-info-settings-btnEl' and contains(@class, 'x-btn-button')]"
        settings_button = wait.until(EC.element_to_be_clickable((By.XPATH, settings_button_xpath)))
        settings_button.click()
        
        logout_option_xpath = "//a[contains(@class, 'x-menu-item-link') and .//span[normalize-space(text())='Esci']]"
        logout_option = wait.until(EC.element_to_be_clickable((By.XPATH, logout_option_xpath)))
        logout_option.click()
        
        time.sleep(1) 
        
        yes_button_xpath = "//a[contains(@class, 'x-btn') and .//span[normalize-space(text())='Si']]"
        yes_button = wait.until(EC.element_to_be_clickable((By.XPATH, yes_button_xpath)))
        driver.execute_script("arguments[0].click();", yes_button)
        logger.info("Logout eseguito.")
        
        WebDriverWait(driver, 10).until(EC.url_contains("Login"))
        logger.info("Ritorno alla pagina di login confermato.")
    except Exception as e_logout:
        logger.warning(f"ATTENZIONE: Errore o timeout durante il logout: {e_logout}")

except Exception as e_selenium_general:
    logger.critical(f"ERRORE GENERALE durante le operazioni Selenium: {e_selenium_general}\n{traceback.format_exc()}")
finally:
    if driver:
        logger.info("-" * 40)
        logger.info("Chiusura del browser WebDriver...")
        driver.quit()
        logger.info("Browser chiuso.")

# --- Sezione 3: Elaborazione File Excel (Invariata) ---
if final_downloaded_path and final_downloaded_path.exists():
    logger.info("\n" + "-" * 50)
    logger.info("INIZIO ELABORAZIONE FILE EXCEL")
    logger.info("-" * 50)
    
    try:
        if not DATABASE_FILE_PATH.exists():
            raise FileNotFoundError(f"File database non trovato: {DATABASE_FILE_PATH}")

        logger.info(f"Apertura file scaricato: {final_downloaded_path.name}")
        wb_source = openpyxl.load_workbook(final_downloaded_path)
        sheet_source = wb_source.active

        logger.info(f"Apertura file database: {DATABASE_FILE_PATH.name}")
        wb_dest = openpyxl.load_workbook(DATABASE_FILE_PATH, keep_vba=True)
        sheet_dest = wb_dest[DATABASE_SHEET_NAME] if DATABASE_SHEET_NAME in wb_dest.sheetnames else wb_dest.create_sheet(DATABASE_SHEET_NAME)

        logger.info("  Indicizzazione righe esistenti nel database (con normalizzazione)...")
        existing_rows = set()
        for row in sheet_dest.iter_rows(min_row=2, values_only=True):
            normalized_row = tuple(str(cell).strip() if cell is not None else "" for cell in row)
            existing_rows.add(normalized_row)
        logger.info(f"  Trovate {len(existing_rows)} righe uniche normalizzate esistenti.")

        start_row = 2
        if sheet_dest.max_row <= 1:
            logger.info("  Il database è vuoto o contiene solo l'header. L'intestazione del nuovo file verrà copiata.")
            start_row = 1
        else:
            logger.info("  Il database contiene dati. Verrà saltata l'intestazione del file scaricato.")

        rows_added = 0
        rows_skipped = 0
        
        for row_index, row_values in enumerate(sheet_source.iter_rows(values_only=True), 1):
            if row_index < start_row:
                continue 
            if not any(cell is not None for cell in row_values):
                continue
            normalized_new_row = tuple(str(cell).strip() if cell is not None else "" for cell in row_values)
            if normalized_new_row not in existing_rows:
                sheet_dest.append(row_values)
                existing_rows.add(normalized_new_row)
                rows_added += 1
            else:
                rows_skipped += 1
        
        logger.info("-" * 20)
        logger.info("  RIEPILOGO PROCESSO:")
        logger.info(f"  - Righe Nuove Aggiunte: {rows_added}")
        logger.info(f"  - Righe Duplicate Saltate: {rows_skipped}")
        logger.info("-" * 20)

        if rows_added > 0:
            logger.info("  Salvataggio delle modifiche sul file database...")
            wb_dest.save(DATABASE_FILE_PATH)
            logger.info("  Salvataggio completato.")
        else:
            logger.info("  Nessuna nuova riga da aggiungere. Il database non è stato modificato.")

        wb_source.close()
        wb_dest.close()

        # Retry logic per eliminazione file temporaneo
        removed = False
        for attempt in range(5):
            try:
                time.sleep(1) # Pausa preventiva
                os.remove(final_downloaded_path)
                logger.info(f"  File temporaneo '{final_downloaded_path.name}' eliminato con successo.")
                removed = True
                break
            except OSError as e_remove:
                logger.warning(f"  ATTENZIONE: File bloccato ({e_remove}). Riprovo eliminazione ({attempt+1}/5)...")
        
        if not removed:
             logger.error(f"  ERRORE: Impossibile eliminare il file scaricato '{final_downloaded_path.name}' dopo vari tentativi.")
            
    except Exception as e_excel_processing:
        logger.critical(f"ERRORE CRITICO durante l'elaborazione dei file Excel: {e_excel_processing}\n{traceback.format_exc()}")
else:
    logger.info("\nNessun file scaricato da processare o download fallito. Lo script termina.")

logger.info("\nScript Python terminato.")