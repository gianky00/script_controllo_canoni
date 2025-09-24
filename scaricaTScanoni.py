from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import time
import traceback
import sys
import os
from pathlib import Path
import shutil
import gc
import logging # <-- AGGIUNTO

# --- CONFIGURAZIONE LOGGING (AGGIUNTO) ---
# Un sistema di logging è più flessibile e informativo dei semplici 'print'.
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)-8s - %(message)s", handlers=[logging.StreamHandler()])
logger = logging.getLogger(__name__)

# Import per pywin32 e definizione di PYWIN32_AVAILABLE
try:
    import win32com.client
    PYWIN32_AVAILABLE = True
except ImportError:
    PYWIN32_AVAILABLE = False
    logger.warning("Libreria pywin32 non è installata. L'esecuzione di macro Excel non sarà possibile.")

# --- NUOVA FUNZIONE DI ATTESA ROBUSTA (AGGIUNTA E ADATTATA) ---
def attendi_scomparsa_overlay(driver, timeout_secondi=45):
    """
    Attende in modo robusto che gli overlay di caricamento tipici dei siti Ext JS scompaiano.
    Rileva maschere di caricamento o messaggi specifici come 'Caricamento...'.
    """
    try:
        overlay_wait = WebDriverWait(driver, timeout_secondi)
        # XPath per maschere di caricamento comuni in Ext JS
        xpath_mask = "//div[contains(@class, 'x-mask-msg') or contains(@class, 'x-mask')][not(contains(@style,'display: none'))]"
        # XPath per il testo di caricamento
        xpath_text = "//div[text()='Caricamento...']"

        # Attende che gli elementi di cui sopra diventino invisibili
        overlay_wait.until(EC.invisibility_of_element_located((By.XPATH, f"{xpath_mask} | {xpath_text}")))
        logger.info(" -> Overlay di caricamento scomparso.")
        time.sleep(0.3) # Piccola pausa per stabilizzare l'interfaccia dopo la scomparsa dell'overlay
        return True
    except TimeoutException:
        logger.warning(f"Timeout ({timeout_secondi}s) durante l'attesa della scomparsa dell'overlay. Proseguo con cautela.")
        return False


# --- Variabili Globali Configurabili ---
LOGIN_URL = "https://portalefornitori.isab.com/Ui/"
DATA_DA_INSERIRE = "01.01.2025"
FORNITORE_DA_SELEZIONARE = "KK10608 - COEMI S.R.L."

# --- Dettagli File Excel e Celle per Percorsi Dinamici ---
SCRIPT_DIRECTORY = Path(__file__).resolve().parent
EXCEL_FILENAME = "parametriScaricoTS.xlsm"
EXCEL_FILE_PATH = SCRIPT_DIRECTORY / EXCEL_FILENAME

SHEET_NAME = "parametri"
USERNAME_CELL = "A3"
PASSWORD_CELL = "B3"
DOWNLOAD_DIR_CELL = "E2"
DIR_SPOSTAMENTO_TS_CELL = "E3"
PERCORSO_FILE_MACRO_CELL = "E4"
CONDIZIONE_MACRO_CELL = "F4"

COL_NUMERO_ODA = "A"
COL_POSIZIONE_ODA = "B"
START_ROW_LOOP = 7
END_ROW_LOOP = 20

# --- Inizializzazione Variabili di Configurazione ---
USERNAME = None
PASSWORD = None
DOWNLOAD_DIR = None
DIR_SPOSTAMENTO_TS = None
PERCORSO_FILE_MACRO = None
ESEGUIRE_MACRO = "NO"

logger.info(f"Tentativo di leggere i dati di configurazione dal file Excel: {EXCEL_FILE_PATH}")
try:
    if not EXCEL_FILE_PATH.exists():
        logger.critical(f"Errore FATALE: File Excel di configurazione non trovato: {EXCEL_FILE_PATH}"); sys.exit("File Excel di configurazione mancante.")

    workbook_config = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    if SHEET_NAME in workbook_config.sheetnames:
        sheet_config = workbook_config[SHEET_NAME]
        USERNAME = sheet_config[USERNAME_CELL].value
        PASSWORD = sheet_config[PASSWORD_CELL].value

        download_dir_excel = sheet_config[DOWNLOAD_DIR_CELL].value
        if download_dir_excel and isinstance(download_dir_excel, str) and download_dir_excel.strip():
            DOWNLOAD_DIR = download_dir_excel.strip()
        else:
            logger.critical(f"Errore FATALE: Percorso DOWNLOAD_DIR non valido o vuoto dalla cella {DOWNLOAD_DIR_CELL}. Valore: '{download_dir_excel}'"); sys.exit("Percorso download mancante.")

        DIR_SPOSTAMENTO_TS = sheet_config[DIR_SPOSTAMENTO_TS_CELL].value
        if not (DIR_SPOSTAMENTO_TS and isinstance(DIR_SPOSTAMENTO_TS, str) and DIR_SPOSTAMENTO_TS.strip()):
            logger.critical(f"Errore FATALE: Percorso DIR_SPOSTAMENTO_TS non valido o vuoto dalla cella {DIR_SPOSTAMENTO_TS_CELL}. Valore: '{DIR_SPOSTAMENTO_TS}'"); sys.exit("Percorso destinazione mancante.")

        if not (USERNAME and PASSWORD):
            logger.critical("Errore FATALE: Username o Password non trovati nel file Excel."); sys.exit("Credenziali mancanti.")

        percorso_file_macro_temp = sheet_config[PERCORSO_FILE_MACRO_CELL].value
        if percorso_file_macro_temp and isinstance(percorso_file_macro_temp, str) and percorso_file_macro_temp.strip():
            PERCORSO_FILE_MACRO = percorso_file_macro_temp.strip()

        eseguire_macro_temp = sheet_config[CONDIZIONE_MACRO_CELL].value
        if isinstance(eseguire_macro_temp, str) and eseguire_macro_temp.strip().upper() in ["SI", "SÌ", "YES"]:
            ESEGUIRE_MACRO = "SI"
        elif isinstance(eseguire_macro_temp, str) and eseguire_macro_temp.strip().upper() == "NO":
            ESEGUIRE_MACRO = "NO"

        DIR_SPOSTAMENTO_TS = DIR_SPOSTAMENTO_TS.strip()
        logger.info("Configurazione letta con successo.")
        logger.info(f"  DOWNLOAD_DIR: '{DOWNLOAD_DIR}'")
        logger.info(f"  DIR_SPOSTAMENTO_TS: '{DIR_SPOSTAMENTO_TS}'")
        logger.info(f"  PERCORSO_FILE_MACRO (da {PERCORSO_FILE_MACRO_CELL}): '{PERCORSO_FILE_MACRO}'")
        logger.info(f"  CONDIZIONE_MACRO (da {CONDIZIONE_MACRO_CELL}): '{ESEGUIRE_MACRO}'")
    else:
        logger.critical(f"Errore FATALE: Foglio '{SHEET_NAME}' non trovato nel file Excel."); sys.exit(f"Foglio '{SHEET_NAME}' non trovato.")
    workbook_config.close()
    logger.info("Workbook di configurazione chiuso.")
except FileNotFoundError:
    logger.critical(f"Errore FATALE: File Excel di configurazione non trovato: {EXCEL_FILE_PATH}"); sys.exit("File Excel non trovato.")
except Exception as e_excel_config:
    logger.critical(f"Errore FATALE durante la lettura della configurazione da Excel: {e_excel_config}\n{traceback.format_exc()}"); sys.exit("Errore configurazione Excel.")

logger.info("\nAvvio script automatico per operazioni web...")
driver = None

try: # TRY PRINCIPALE per le operazioni Selenium
    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "safeBrowse.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_argument("--disable-gpu"); chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage"); chrome_options.add_argument("--start-maximized")

    logger.info(f"Inizializzazione WebDriver Chrome...")
    driver = webdriver.Chrome(options=chrome_options)

    wait = WebDriverWait(driver, 20)
    popup_wait = WebDriverWait(driver, 7)
    long_wait_for_dropdown_elements = WebDriverWait(driver, 15)

    logger.info(f"Navigazione a: {LOGIN_URL}")
    driver.get(LOGIN_URL)

    logger.info("Tentativo di login...")
    wait.until(EC.presence_of_element_located((By.NAME, "Username"))).send_keys(USERNAME)
    wait.until(EC.presence_of_element_located((By.NAME, "Password"))).send_keys(PASSWORD)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Accedi' and contains(@class, 'x-btn-inner')]"))).click()
    logger.info("Login effettuato. Attendo scomparsa overlay...")
    
    # --- MODIFICA: Sostituzione di time.sleep con attesa robusta ---
    attendi_scomparsa_overlay(driver, 60) # Attende fino a 60 secondi dopo il login

    logger.info("Controllo eventuale pop-up post-login...")
    try:
        ok_button_popup = popup_wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='OK' and contains(@class, 'x-btn-inner')]")))
        logger.info("Pop-up 'OK' trovato. Click...")
        ok_button_popup.click()
        WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.XPATH, "//span[text()='OK' and contains(@class, 'x-btn-inner')]")))
        logger.info("Popup gestito.")
    except TimeoutException:
        logger.info("Nessun pop-up 'OK' rilevato (normale se non previsto).")

    logger.info("Attesa caricamento pagina principale e navigazione menu...")
    
    logger.info("Click su 'Report'...")
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[normalize-space(text())='Report']"))).click()
        logger.info("'Report' cliccato.")
        attendi_scomparsa_overlay(driver) # <-- MODIFICA
    except Exception as e_report_click:
        logger.critical(f"Errore FATALE durante il click su 'Report': {e_report_click}\n{traceback.format_exc()}"); sys.exit("Fallimento navigazione menu (Report).")

    logger.info("Click su 'Timesheet'...")
    fornitore_arrow_xpath = "//div[starts-with(@id, 'generic_refresh_combo_box-') and contains(@id, '-trigger-picker') and contains(@class, 'x-form-arrow-trigger')]"
    try:
        timesheet_menu_xpath = "//span[contains(@id, 'generic_menu_button-') and contains(@id, '-btnEl')][.//span[text()='Timesheet']]"
        wait.until(EC.element_to_be_clickable((By.XPATH, timesheet_menu_xpath))).click()
        logger.info("'Timesheet' cliccato.")
        wait.until(EC.visibility_of_element_located((By.XPATH, fornitore_arrow_xpath)))
        logger.info("Pagina Timesheet (con dropdown Fornitore) caricata.")
        attendi_scomparsa_overlay(driver) # <-- MODIFICA
    except Exception as e_timesheet_click:
        logger.critical(f"Errore FATALE durante il click su 'Timesheet': {e_timesheet_click}\n{traceback.format_exc()}"); sys.exit("Fallimento navigazione menu (Timesheet).")

    logger.info("Impostazione Fornitore e Data Da (operazioni eseguite una sola volta)...")
    try:
        logger.info(f"  Tentativo di selezionare il fornitore: '{FORNITORE_DA_SELEZIONARE}'...")
        fornitore_arrow_element = wait.until(EC.element_to_be_clickable((By.XPATH, fornitore_arrow_xpath)))
        ActionChains(driver).move_to_element(fornitore_arrow_element).click().perform()
        logger.info(f"  Click sulla freccia del dropdown Fornitore eseguito.")

        coemi_srl_option_xpath = f"//li[normalize-space(text())='{FORNITORE_DA_SELEZIONARE}']"
        coemi_srl_option = long_wait_for_dropdown_elements.until(EC.presence_of_element_located((By.XPATH, coemi_srl_option_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'nearest'});", coemi_srl_option)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", coemi_srl_option)
        logger.info(f"  Fornitore '{FORNITORE_DA_SELEZIONARE}' selezionato tramite JS.")
        
        # --- MODIFICA: Sostituzione di time.sleep con attesa robusta ---
        attendi_scomparsa_overlay(driver) # Attende dopo la selezione del fornitore

        logger.info(f"  Inserimento data '{DATA_DA_INSERIRE}'...")
        campo_data_da = wait.until(EC.visibility_of_element_located((By.NAME, "DataTimesheetDa")))
        campo_data_da.clear()
        campo_data_da.send_keys(DATA_DA_INSERIRE)
        logger.info(f"  Data '{DATA_DA_INSERIRE}' inserita.")
        logger.info("Impostazioni fisse (Fornitore, Data) completate.")
    except Exception as e_setup_fisso:
        logger.critical(f"Errore FATALE durante le impostazioni fisse (Fornitore/Data Da): {e_setup_fisso}\n{traceback.format_exc()}"); sys.exit("Fallimento impostazioni fisse.")

    js_dispatch_events_corretto = """
        var el = arguments[0]; var ev_in = new Event('input', {bubbles:true}); el.dispatchEvent(ev_in);
        var ev_ch = new Event('change', {bubbles:true}); el.dispatchEvent(ev_ch);"""

    logger.info(f"\nApertura workbook per loop dati: {EXCEL_FILE_PATH}")
    workbook_loop_obj = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheet_dati_loop = workbook_loop_obj[SHEET_NAME]

    for riga_corrente in range(START_ROW_LOOP, END_ROW_LOOP + 1):
        logger.info("-" * 40)
        logger.info(f"Processo la riga {riga_corrente} del file Excel...")

        numero_oda_excel_loop = str(sheet_dati_loop[f"{COL_NUMERO_ODA}{riga_corrente}"].value or "").strip()
        if not numero_oda_excel_loop:
            logger.info(f"Cella {COL_NUMERO_ODA}{riga_corrente} vuota. Loop terminato."); break
        
        posizione_oda_excel_loop = str(sheet_dati_loop[f"{COL_POSIZIONE_ODA}{riga_corrente}"].value or "").strip()
        if not posizione_oda_excel_loop:
            logger.info(f"Info: Cella {COL_POSIZIONE_ODA}{riga_corrente} vuota. Si procede con il solo Numero OdA.")

        logger.info(f"  Dati letti: NumeroOdA='{numero_oda_excel_loop}', PosizioneOdA='{posizione_oda_excel_loop}'")

        try:
            campo_numero_oda = wait.until(EC.presence_of_element_located((By.NAME, "NumeroOda")))
            driver.execute_script("arguments[0].value = arguments[1];", campo_numero_oda, numero_oda_excel_loop)
            driver.execute_script(js_dispatch_events_corretto, campo_numero_oda)
            logger.info(f"  Valore '{numero_oda_excel_loop}' impostato in 'NumeroOda'.")

            campo_posizione_oda = wait.until(EC.presence_of_element_located((By.NAME, "PosizioneOda")))
            driver.execute_script("arguments[0].value = '';", campo_posizione_oda)
            driver.execute_script("arguments[0].value = arguments[1];", campo_posizione_oda, posizione_oda_excel_loop)
            driver.execute_script(js_dispatch_events_corretto, campo_posizione_oda)
            logger.info(f"  Valore '{posizione_oda_excel_loop}' impostato in 'PosizioneOda'.")

            pulsante_cerca_xpath = "//a[contains(@class, 'x-btn') and @role='button'][.//span[normalize-space(text())='Cerca' and contains(@class, 'x-btn-inner')]]"
            wait.until(EC.element_to_be_clickable((By.XPATH, pulsante_cerca_xpath))).click()
            logger.info("  Pulsante 'Cerca' cliccato. Attesa risultati...")

            # --- MODIFICA: Sostituzione di time.sleep con attesa robusta ---
            attendi_scomparsa_overlay(driver, 90) # Attende fino a 90 secondi per i risultati della ricerca

            logger.info("  Tentativo di download del file Excel...")
            path_to_downloads_obj = Path(DOWNLOAD_DIR)
            files_before_download = {f for f in path_to_downloads_obj.iterdir() if f.is_file() and f.suffix.lower() == '.xlsx'}

            excel_button_xpath = "//div[contains(@class, 'x-tool') and @role='button'][.//div[@data-ref='toolEl' and contains(@class, 'x-tool-tool-el') and contains(@style, 'FontAwesome')]]"
            wait.until(EC.element_to_be_clickable((By.XPATH, excel_button_xpath))).click()
            logger.info("  Icona Excel per download cliccata. Attesa download (max 25s)...")

            downloaded_file_path = None; download_start_time = time.time()
            while time.time() - download_start_time < 25:
                current_files = {f for f in path_to_downloads_obj.iterdir() if f.is_file() and f.suffix.lower() == '.xlsx'}
                new_files = current_files - files_before_download
                if new_files:
                    downloaded_file_path = max(list(new_files), key=lambda f: f.stat().st_mtime)
                    logger.info(f"  File scaricato rilevato: {downloaded_file_path.name}"); break
                time.sleep(0.5)

            if not downloaded_file_path:
                logger.warning("  File non rilevato rapidamente, controllo finale dopo breve attesa..."); time.sleep(5)
                current_files = {f for f in path_to_downloads_obj.iterdir() if f.is_file() and f.suffix.lower() == '.xlsx'}
                new_files = current_files - files_before_download
                if new_files: downloaded_file_path = max(list(new_files), key=lambda f: f.stat().st_mtime)

            if downloaded_file_path and downloaded_file_path.exists():
                nome_base_pos = f"-{posizione_oda_excel_loop}" if posizione_oda_excel_loop else ""
                nuovo_nome_base = f"{numero_oda_excel_loop}{nome_base_pos}"
                nuovo_nome_file = f"{nuovo_nome_base}.xlsx"
                percorso_file_rinominato_download = path_to_downloads_obj / nuovo_nome_file

                counter = 1
                while percorso_file_rinominato_download.exists() and percorso_file_rinominato_download.resolve() != downloaded_file_path.resolve():
                    timestamp_attuale = time.strftime("%Y%m%d-%H%M%S")
                    nuovo_nome_file = f"{nuovo_nome_base}-{timestamp_attuale}_{counter}.xlsx"
                    percorso_file_rinominato_download = path_to_downloads_obj / nuovo_nome_file
                    counter += 1

                downloaded_file_path.rename(percorso_file_rinominato_download)
                logger.info(f"  File rinominato in directory Downloads: {percorso_file_rinominato_download.name}")

                path_destinazione_finale = Path(DIR_SPOSTAMENTO_TS)
                path_destinazione_finale.mkdir(parents=True, exist_ok=True)
                file_destinazione_con_nome_univoco = path_destinazione_finale / percorso_file_rinominato_download.name

                shutil.move(str(percorso_file_rinominato_download), str(file_destinazione_con_nome_univoco))
                logger.info(f"  File spostato con successo in: {file_destinazione_con_nome_univoco}")
            else:
                logger.error("  ERRORE CRITICO: Download fallito o file non trovato dopo attesa. Riga saltata."); continue

        except Exception as e_loop_riga:
            logger.error(f"  Errore durante il processo della riga {riga_corrente}: {e_loop_riga}\n{traceback.format_exc()}"); continue

        logger.info(f"Riga {riga_corrente} processata con successo."); time.sleep(1)

    logger.info("\nLoop di processamento righe Excel terminato.")
    workbook_loop_obj.close()
    logger.info("Workbook dati del loop chiuso.")

    logger.info("-" * 40)
    logger.info("Tentativo di Logout...")
    try:
        settings_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@id='user-info-settings-btnEl' and contains(@class, 'x-btn-button')]")))
        settings_button.click()
        logger.info("Pulsante Settings (ingranaggio) cliccato.")

        logout_option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(@class, 'x-menu-item-link')][.//span[normalize-space(text())='Esci']]")))
        logout_option.click()
        logger.info("Opzione 'Esci' cliccata.")

        time.sleep(2) 

        yes_button_xpath = "//a[contains(@class, 'x-btn') and @role='button'][.//span[normalize-space(text())='Si']]"
        yes_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, yes_button_xpath)))
        logger.info("Pulsante 'Si' per conferma logout trovato nel DOM.")
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", yes_button)
        logger.info("Logout (click su 'Si' via JS) eseguito.")

        WebDriverWait(driver, 10).until(EC.url_contains(LOGIN_URL.split("://")[1].split("/")[0]))
        logger.info(f"URL dopo logout: {driver.current_url} - Ritorno alla pagina di login confermato.")
    except TimeoutException:
        current_url_after_logout = driver.current_url if driver else "N/A"
        logger.warning(f"Timeout durante il logout o la conferma. URL attuale: {current_url_after_logout}.")
        if driver:
            try:
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.NAME, "Username")))
                logger.info("Campo Username trovato. Logout probabilmente riuscito nonostante il timeout sull'URL.")
            except TimeoutException:
                logger.warning("ATTENZIONE: Logout incerto. Né l'URL atteso né il campo Username sono stati confermati rapidamente.")
    except Exception as e_logout:
        logger.error(f"Errore durante il tentativo di Logout: {e_logout}\n{traceback.format_exc()}")

    logger.info("Tutte le operazioni web (incluso il tentativo di logout) sono state completate!")

except Exception as e_selenium_general:
    logger.critical(f"Errore generale durante le operazioni Selenium: {e_selenium_general}\n{traceback.format_exc()}")
finally:
    if driver:
        logger.info("-" * 40)
        secondi_di_pausa_prima_chiusura = 3
        logger.info(f"Pausa di {secondi_di_pausa_prima_chiusura} secondi prima di chiudere il browser...")
        time.sleep(secondi_di_pausa_prima_chiusura)

        logger.info("Chiusura del browser WebDriver in corso...")
        try:
            driver.quit()
            logger.info("Browser WebDriver chiuso con successo.")
            driver = None
        except Exception as e_quit_driver:
            logger.error(f"Errore durante la chiusura del browser WebDriver: {e_quit_driver}")

if ESEGUIRE_MACRO == "SI":
    # [La sezione MACRO rimane invariata]
    logger.info("-" * 40)
    logger.info("Inizio sezione esecuzione macro Excel...")
    if not PYWIN32_AVAILABLE:
        logger.critical("ERRORE FATALE: La libreria pywin32 non è installata. Impossibile eseguire la macro Excel.")
    elif not (PERCORSO_FILE_MACRO and Path(PERCORSO_FILE_MACRO).is_file()):
        logger.error(f"ERRORE: Condizione 'SI' per macro, ma il file specificato non è valido o non trovato.")
        logger.error(f"Percorso letto da Excel ({PERCORSO_FILE_MACRO_CELL}): '{PERCORSO_FILE_MACRO}'")
    else:
        logger.info(f"Condizione 'SI' trovata. Tento di eseguire la macro 'elaboraTutto()' nel file: {PERCORSO_FILE_MACRO}")
        logger.info("\n*** ATTENZIONE: L'applicazione Excel verrà resa VISIBILE per il debug. ***")
        logger.info("    Osserva attentamente l'istanza di Excel che si apre.")
        logger.info("    Cerca eventuali messaggi di errore o comportamenti inattesi mostrati direttamente da Excel.\n")

        excel_app_macro, workbook_macro_to_run = None, None
        try:
            abs_path_macro_file = str(Path(PERCORSO_FILE_MACRO).resolve())

            logger.info(f"Tentativo di avviare Excel.Application e aprire il workbook: {abs_path_macro_file}")
            excel_app_macro = win32com.client.Dispatch("Excel.Application")
            excel_app_macro.Visible = True 
            excel_app_macro.DisplayAlerts = True 

            workbook_macro_to_run = excel_app_macro.Workbooks.Open(abs_path_macro_file)
            logger.info(f"Workbook '{Path(PERCORSO_FILE_MACRO).name}' aperto con successo.")

            logger.info("Esecuzione della macro 'elaboraTutto' in corso...")
            excel_app_macro.Run("elaboraTutto") 
            logger.info("Chiamata alla macro 'elaboraTutto()' completata. Controlla l'istanza di Excel.")

            input("\n--- ISPEZIONA EXCEL ORA --- \nPremi INVIO qui nel terminale per procedere con la chiusura di Excel...")

        except Exception as e_macro_exec:
            logger.critical(f"ERRORE CRITICO durante l'esecuzione della macro Excel: {e_macro_exec}\n{traceback.format_exc()}")
            if excel_app_macro:
                input("\n--- ERRORE DURANTE MACRO --- \nExcel potrebbe essere aperto. Premi INVIO per tentare la chiusura...")
        finally:
            if workbook_macro_to_run is not None:
                try:
                    workbook_macro_to_run.Close(SaveChanges=True)
                    logger.info("Workbook contenente la macro chiuso.")
                except Exception as e_close_wb_macro:
                    logger.warning(f"Attenzione: Errore chiusura workbook macro: {e_close_wb_macro}")

            if excel_app_macro is not None:
                try:
                    excel_app_macro.Quit()
                    logger.info("Applicazione Excel chiusa.")
                except Exception as e_quit_excel_app:
                    logger.warning(f"Attenzione: Errore chiusura Excel: {e_quit_excel_app}")

            if workbook_macro_to_run is not None: del workbook_macro_to_run
            if excel_app_macro is not None: del excel_app_macro
            workbook_macro_to_run, excel_app_macro = None, None 

            gc.collect() 
            logger.info("Pulizia oggetti COM e garbage collection completata.")
            time.sleep(1) 

elif ESEGUIRE_MACRO == "NO":
    logger.info("-" * 40)
    logger.info("Condizione 'NO' trovata. Nessuna macro Excel da eseguire.")
else:
    logger.info("-" * 40)
    logger.info(f"Valore '{ESEGUIRE_MACRO}' non riconosciuto per condizione macro. Nessuna macro eseguita.")

logger.info("\nScript Python terminato.")
